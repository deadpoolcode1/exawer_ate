"""Write a Plan to xlsx — DHCP-snoopy 9-column shape (M1 client respin).

Layout (matches references/DHCP-snoopy_TP_with_PW.xlsx):

  - **Topic / Action / Req ID / Expectation / Monitor / Equipment /
    Build / Results / Comment** — 9 columns, atomic-row-under-topic-banner.
  - The generator still produces multi-line Setup/Action/Verify PlanRow
    blobs (so the AI-enrichment cache survives the shape change);
    `atomic_rows.rows_for_plan_row()` decomposes each blob into a banner
    row + N atomic action rows at render time.
  - RFC mandates that no flow claims are emitted by `generator.py` as
    first-class PlanRows (Yossi push-back 2026-05-21) — they flow
    through the enricher and render on the main sheet alongside flow
    rows, tinted green as "RFC mandate" so QA still sees the source.
  - **Coverage sheet** unchanged from the previous respin; reports
    requirement → flow coverage with the same orphan-highlight semantics
    as before.
  - **Feature Concept Catalog** stays at the top of the main sheet.

We do not modify the customer's original template file.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ate.planner.atomic_rows import AtomicRow, rows_for_plan_row
from ate.planner.feature_catalog import build_catalog
from ate.planner.flows import EVPN_FLOWS
from ate.planner.model import Plan, PlanRow, Requirement

# DHCP-snoopy 9-column schema. Headers chosen to match the reference TP
# exactly where possible (Topic / Action / Expectation / Monitor / Build /
# Results / Comment); Req ID and Equipment are added so traceability
# survives the shape change.
HEADER_ROW = [
    "Topic",
    "Action",
    "SFS / RFC Req ID",
    "Expectation",
    "Monitor (show / verify command)",
    "Test Equipment",
    "Build number",
    "Results (Pass\\Fail)",
    "Comment \\ Bug number",
]

FLOW_FILL = PatternFill("solid", fgColor="D7E4F4")
CAT_FILL = PatternFill("solid", fgColor="E9ECEF")
SUBCAT_FILL = PatternFill("solid", fgColor="F8F9FA")
HEADER_FILL = PatternFill("solid", fgColor="343A40")
CATALOG_FILL = PatternFill("solid", fgColor="DCE9F1")
CATALOG_GROUP_FILL = PatternFill("solid", fgColor="B6CCDC")
ORPHAN_FILL = PatternFill("solid", fgColor="F9E2D6")
HEADER_FONT = Font(color="FFFFFF", bold=True)
META_FONT = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_LEFT = Alignment(wrap_text=True, vertical="top", horizontal="left")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def _row_height_for(text: str, width: int = 70) -> float:
    """Estimate the row height needed to show `text` in full at column
    `width` (≈ chars per line). Cap raised to 320 pt (well under Excel's
    409 max) so cells render their entire content instead of truncating
    — client 2026-06-01, item 1.
    """
    if not text:
        return 18
    per_line = max(1, int(width) - 2)
    lines = text.count("\n") + 1
    extra = sum(max(0, len(ln) // per_line) for ln in text.split("\n"))
    return min(18 + 14 * (lines + extra - 1), 320)


def _row_height_multi(cells: list[tuple[str, int]]) -> float:
    """Tallest height across several (text, column-width) cells — so a row
    is sized to its longest column, not just one. Fixes the truncation
    where height tracked the Action cell while Expectation / Monitor
    overflowed (client 2026-06-01, item 1)."""
    return max((_row_height_for(t, w) for t, w in cells if t), default=18)


def _write_catalog(ws, catalog: list, start_row: int) -> int:
    row = start_row
    cell = ws.cell(row=row, column=1, value="Feature Concept Catalog")
    cell.font = Font(bold=True, size=12, color="1F3A5F")
    row += 1
    cell = ws.cell(
        row=row, column=1,
        value=("EVPN concepts the plan covers — values sourced from the "
               "EVPN CLI doc and RFC 7432bis. Reviewers can map any plan "
               "row to its concept here."),
    )
    cell.alignment = WRAP_LEFT
    cell.font = Font(italic=True, color="555555")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    row += 1
    for c, label in enumerate(("Concept", "Value", "Description / Source"), 1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.font = HEADER_FONT
        cell.fill = CATALOG_GROUP_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 22
    row += 1
    for entry in catalog:
        cell = ws.cell(row=row, column=1, value=entry.name)
        cell.font = META_FONT
        cell.fill = CATALOG_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = THIN_BORDER
        cell2 = ws.cell(row=row, column=2, value="(source)")
        cell2.fill = CATALOG_FILL
        cell2.font = Font(italic=True, color="555555")
        cell2.border = THIN_BORDER
        cell3 = ws.cell(row=row, column=3, value=entry.source)
        cell3.fill = CATALOG_FILL
        cell3.alignment = WRAP_LEFT
        cell3.font = Font(italic=True, color="555555")
        cell3.border = THIN_BORDER
        ws.row_dimensions[row].height = 20
        row += 1
        for value, description in entry.values:
            ws.cell(row=row, column=1).border = THIN_BORDER
            ws.cell(row=row, column=2, value=value).alignment = WRAP_LEFT
            ws.cell(row=row, column=2).border = THIN_BORDER
            ws.cell(row=row, column=3, value=description).alignment = WRAP_LEFT
            ws.cell(row=row, column=3).border = THIN_BORDER
            ws.row_dimensions[row].height = _row_height_for(description)
            row += 1
        if entry.notes:
            cell = ws.cell(row=row, column=1, value="(notes)")
            cell.font = Font(italic=True, color="555555")
            cell.border = THIN_BORDER
            ws.cell(row=row, column=2).border = THIN_BORDER
            cell3 = ws.cell(row=row, column=3, value=entry.notes)
            cell3.alignment = WRAP_LEFT
            cell3.font = Font(italic=True, color="555555")
            cell3.border = THIN_BORDER
            ws.row_dimensions[row].height = _row_height_for(entry.notes)
            row += 1
    return row


RFC_FILL = PatternFill("solid", fgColor="D4EDDA")      # RFC mandate row (normative)
INHERIT_FILL = PatternFill("solid", fgColor="E2D9F3")  # CLI-inheritance row

# ── Section bands — DHCP-snoopy "division to topics" (client 2026-05-25) ──
# `references/DHCP-snoopy_TP_with_PW.xlsx` groups its topic banners under a
# top tier of functional section headers (CLI configuration → clear → show
# → Feature Functionality → Feature interaction → non-functional). We mirror
# that here: every PlanRow is assigned a band, the body is stable-sorted by
# band (within-band order preserved, so CLI commands stay in catalog order
# and flows stay in FLOW-NNN order), and a dark section header is emitted at
# each band transition. CLI-first ordering per client direction.
SEC_CLI_CONFIG = "CLI Configuration"
SEC_CLEAR = "CLI Clear"
SEC_SHOW = "Show Commands"
SEC_FUNC = "Feature Functionality"
SEC_NONFUNC = "Feature Interaction, Scale & Lifecycle"
SEC_RFC = "RFC Protocol Mandates"
SEC_OTHER = "Additional Tests"

# Section band order. CLI configuration rows now carry their *functional
# group* directly in PlanRow.category (client 2026-06-02, Eyal Ozeri, item
# 2: interface / LACP / l2-EVPN / l2-VPLS / BGP-AF grouped), and show rows
# carry a new-vs-modified category (item: "differentiation between new show
# commands and modified show commands"). The order below is matched by
# prefix so the exact suffix of a category label doesn't matter.
_SECTION_RANK_PREFIXES = [
    "CLI Configuration — Interface",
    "CLI Configuration — LACP",
    "CLI Configuration — L2-Services EVPN",
    "CLI Configuration — L2-Services VPLS",
    "CLI Configuration — BGP EVPN",
    "CLI Configuration — Other",
    "CLI Configuration",            # legacy / fallback config band
    "CLI Show — New",
    "CLI Show — Modified",
    "CLI Clear",
    SEC_FUNC,
    SEC_NONFUNC,
    SEC_RFC,
    SEC_OTHER,
]


def _section_rank(label: str) -> int:
    for i, pref in enumerate(_SECTION_RANK_PREFIXES):
        if label.startswith(pref):
            return i
    return len(_SECTION_RANK_PREFIXES)


SECTION_FILL = PatternFill("solid", fgColor="1F3A5F")  # deep navy band
SECTION_FONT = Font(color="FFFFFF", bold=True, size=11)


def _section_for_row(r: PlanRow, flow_lookup: dict) -> str:
    """Assign a PlanRow to a DHCP-snoopy-style section band.

    RFC-mandate rows (no flow, RFC* req id) → RFC band. Flow rows split by
    the flow's `coverage_driven` flag: requirement-anchored flows are the
    functional use cases; coverage-driven flows (scale / upgrade / NETCONF /
    soak / access-variants) are the broad-technique non-functional band.
    CLI rows carry their band label directly in `category` — the functional
    config group, or the show-new / show-modified / clear band (client
    2026-06-02).
    """
    if (not r.flow_id) and r.sfs_requirement_id.startswith("RFC"):
        return SEC_RFC
    if r.flow_id:
        flow = flow_lookup.get(r.flow_id)
        if flow is not None and flow.coverage_driven:
            return SEC_NONFUNC
        return SEC_FUNC
    cat = (r.category or "").strip()
    if cat.startswith("CLI "):
        return cat
    # Legacy fallback (CLI rows authored before the grouping change).
    name = (r.sub_category or "").strip().lower()
    if name.startswith("show"):
        return SEC_SHOW
    if name.startswith("clear"):
        return SEC_CLEAR
    if name:
        return SEC_CLI_CONFIG
    return SEC_OTHER


def _write_section_header(ws, label: str, row: int) -> int:
    """Render a top-tier section band (dark, full width) above its topic
    banners — the visual divider that DHCP-snoopy uses between CLI
    configuration / show / clear / Feature Functionality / … groups."""
    cell = ws.cell(row=row, column=1, value=label.upper())
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = Alignment(wrap_text=True, vertical="center",
                                horizontal="left")
    for c in range(2, len(HEADER_ROW) + 1):
        ws.cell(row=row, column=c).fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1,
                    end_row=row, end_column=len(HEADER_ROW))
    ws.row_dimensions[row].height = 26
    return row + 1


# Kind-priority for picking the most informative marker when a flow row
# aggregates multiple SFS reqs of different kinds. delta wins because a
# vendor-modified clause is the most surprising thing a QA engineer can
# encounter; overlay next; pointer last (it adds the least signal).
_KIND_PRIORITY = ("delta", "overlay", "pointer",
                  "sfs_with_rfc_context", "base_sfs")


def _provenance_for_row(r: PlanRow,
                        req_by_id: dict[str, Requirement],
                        inherited_names: set[str]) -> str:
    """Compute the provenance / kind-marker string for a PlanRow.

    Returns one of:
      "rfc-orphan"          — RFC mandate row (generator-emitted synth)
      "cli-inherit"         — BGP sub-config row
      "delta:<rfc_link>"    — SFS req that modifies an RFC clause
      "overlay:<rfc_link>"  — SFS req that adds beyond an RFC clause
      "pointer:<rfc_link>"  — SFS req that just points at an RFC clause
      ""                    — default (flow row anchored by base SFS)

    The kind markers carry the linked RFC req_id (when known) so the
    Comment column can read "delta from RFC7432bis-§8.5" — a QA engineer
    sees the SFS-vs-RFC relationship without opening the source docs.
    """
    # Source markers first — they win over kind markers because rfc-
    # orphan / cli-inherit rows have their own dedicated visual tier.
    is_inherited = bool(r.sub_category and r.sub_category in inherited_names)
    is_rfc_orphan = (not r.flow_id) and r.sfs_requirement_id.startswith("RFC")
    if is_rfc_orphan:
        return "rfc-orphan"
    if is_inherited:
        return "cli-inherit"

    # Walk covered_req_ids → pick the SFS req with the highest-priority
    # kind. RFC reqs in covered_req_ids are skipped (their relationship
    # is already implicit — "rfc"). Multi-req flow rows often mix SFS
    # and RFC; we want the SFS relationship marker.
    best_kind = ""
    best_link = ""
    best_priority = len(_KIND_PRIORITY)
    for rid in r.covered_req_ids:
        req = req_by_id.get(rid)
        if req is None or req.source != "spec":
            continue
        if req.kind not in _KIND_PRIORITY:
            continue
        priority = _KIND_PRIORITY.index(req.kind)
        if priority < best_priority:
            best_priority = priority
            best_kind = req.kind
            best_link = req.rfc_links[0] if req.rfc_links else ""

    if best_kind in ("delta", "overlay", "pointer"):
        return f"{best_kind}:{best_link}"
    return ""


def _comment_for_provenance(provenance: str) -> str:
    """Map provenance string → human-readable Comment-column marker.

    Yossi 2026-05-21 follow-up: SFS reqs classified as delta / overlay /
    pointer carry the linked RFC req_id in the provenance string
    ("delta:RFC7432bis-§8.5") so a QA engineer can see the SFS-vs-RFC
    relationship at a glance. The two source markers (rfc-orphan,
    cli-inherit) take the same column but never co-occur with kind
    markers since RFC-orphan reqs have kind="rfc" and inherited CLI
    rows have kind="cli".
    """
    if provenance == "rfc-orphan":
        return "RFC mandate"
    if provenance == "cli-inherit":
        return "CLI inheritance"
    if provenance.startswith("delta:"):
        link = provenance.split(":", 1)[1]
        return f"delta from {link}" if link else "delta from RFC base"
    if provenance.startswith("overlay:"):
        link = provenance.split(":", 1)[1]
        return f"overlay on {link}" if link else "overlay on RFC base"
    if provenance.startswith("pointer:"):
        link = provenance.split(":", 1)[1]
        return f"pointer to {link}" if link else "pointer to RFC"
    return ""


def _fill_for_provenance(provenance: str) -> PatternFill | None:
    """Banner / row tint per provenance. Returns None for default (flow)."""
    if provenance == "rfc-orphan":
        return RFC_FILL
    if provenance == "cli-inherit":
        return INHERIT_FILL
    # delta / overlay / pointer rows keep the flow tint — the kind
    # marker in col 9 carries the relationship info without piling on
    # visual noise.
    return None


def _write_atomic_row(ws, ar: AtomicRow, row: int) -> int:
    """Render one AtomicRow into the 9-column schema. Banner rows merge
    A→F so the topic reads like a section header (matches DHCP-snoopy).
    """
    comment = _comment_for_provenance(ar.provenance)
    tint = _fill_for_provenance(ar.provenance)

    if ar.is_banner:
        cell = ws.cell(row=row, column=1, value=ar.topic)
        cell.font = META_FONT
        fill = tint or FLOW_FILL
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        for c in range(2, len(HEADER_ROW) + 1):
            ws.cell(row=row, column=c).fill = fill
        ws.merge_cells(start_row=row, start_column=1,
                        end_row=row, end_column=6)
        if comment:
            ws.cell(row=row, column=9, value=comment)
        ws.row_dimensions[row].height = 22
        return row + 1

    # Continuation row: col A intentionally empty (inherits topic above).
    ws.cell(row=row, column=1, value="").alignment = WRAP_LEFT
    ws.cell(row=row, column=2, value=ar.action).alignment = WRAP_LEFT
    ws.cell(row=row, column=3,
            value=", ".join(ar.req_ids) if ar.req_ids else "")\
        .alignment = WRAP_LEFT
    ws.cell(row=row, column=4, value=ar.expectation).alignment = WRAP_LEFT
    ws.cell(row=row, column=5,
            value=", ".join(ar.monitor) if ar.monitor else "")\
        .alignment = WRAP_LEFT
    ws.cell(row=row, column=6, value=ar.equipment).alignment = WRAP_LEFT
    ws.cell(row=row, column=9, value=comment).alignment = WRAP_LEFT
    # Subtle row tint for source-marker rows (RFC / inherit). Delta /
    # overlay / pointer rows do not tint — the comment marker suffices.
    if tint is not None:
        for c in range(1, len(HEADER_ROW) + 1):
            ws.cell(row=row, column=c).fill = tint
    # Size the row to its tallest column (Action@60, Req@22, Expectation@60,
    # Monitor@36, Equipment@28) so no cell truncates — col widths mirror the
    # `widths` list in write_xlsx.
    ws.row_dimensions[row].height = _row_height_multi([
        (ar.action, 60),
        (", ".join(ar.req_ids), 22),
        (ar.expectation, 60),
        (", ".join(ar.monitor), 36),
        (ar.equipment, 28),
    ])
    return row + 1


def _write_flow_banner(ws, flow_id: str, flow_name: str, summary: str,
                       covered_ids: list[str], row: int) -> int:
    """Render a Flow banner row spanning the data columns. The banner
    states the flow ID + name, a short summary, and the covered req-IDs
    so reviewers can see scope before reading per-category rows.
    """
    label = f"{flow_id} — {flow_name}"
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = META_FONT
    cell.fill = FLOW_FILL
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    for c in range(2, len(HEADER_ROW) + 1):
        ws.cell(row=row, column=c).fill = FLOW_FILL
    ws.merge_cells(start_row=row, start_column=2, end_row=row,
                    end_column=len(HEADER_ROW))
    body = summary
    if covered_ids:
        body += f"\nCovers: {', '.join(covered_ids)}"
    cell2 = ws.cell(row=row, column=2, value=body)
    cell2.alignment = WRAP_LEFT
    cell2.fill = FLOW_FILL
    ws.row_dimensions[row].height = _row_height_for(body)
    return row + 1


# Known orphan requirements that are deliberately out of M1 scope.
# Reviewers see these in the Coverage sheet annotated with the reason
# rather than a bare "(orphan)" — so a 100% strict reading still
# returns a defensible answer.
#
# Yossi 2026-05-21: RFC mandates do not get out-of-scope exemptions any
# more — every RFC clause is promoted to a first-class row by the synth
# path in generator.py. Only SFS-side meta-requirements stay here.
_INTENTIONAL_ORPHANS: dict[str, str] = {
    "EVPNS-REQ#10": (
        "(orphan: META — lists supported RFCs; not a runnable use case)"
    ),
}


def _write_coverage_sheet(wb, plan: Plan) -> None:
    """Render the Coverage sheet: req-id → flows-that-cover-it.

    Orphans (req-IDs no flow claims) are highlighted at the top so the
    reviewer can extend the flow catalog or accept that some
    requirements are covered exclusively by the per-command CLI rows.
    """
    coverage: dict[str, list[str]] = plan.__dict__.get("_coverage", {})
    orphans: list[str] = plan.__dict__.get("_orphans", [])
    ws = wb.create_sheet("Coverage")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 50

    row = 1
    ws.cell(row=row, column=1, value="Requirement → Flow coverage").font = (
        Font(bold=True, size=12, color="1F3A5F")
    )
    row += 1
    cell = ws.cell(
        row=row, column=1,
        value=("Each spec / RFC requirement listed below maps to the flows "
               "(use cases) that exercise it. Requirements highlighted in "
               "orange are not claimed by any flow — the flow catalog "
               "should be extended to cover them, or they are covered "
               "exclusively by the per-command CLI Configuration rows."),
    )
    cell.alignment = WRAP_LEFT
    cell.font = Font(italic=True, color="555555")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.row_dimensions[row].height = _row_height_for(cell.value)
    row += 2

    for c, label in enumerate(
        ("Req ID", "Section", "Title", "Covering Flow IDs"), 1
    ):
        cell = ws.cell(row=row, column=c, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[row].height = 24
    row += 1

    # Stable order: spec reqs first (EVPNS-REQ#NN by numeric N), RFC reqs after.
    reqs_in_order = [r for r in plan.requirements if not r.req_id.startswith("CLI:")]
    for r in reqs_in_order:
        flow_ids = coverage.get(r.req_id, [])
        if flow_ids:
            cov_cell = ", ".join(flow_ids)
        else:
            cov_cell = _INTENTIONAL_ORPHANS.get(r.req_id, "(orphan)")
        cells_row = [
            r.req_id,
            r.section_number or "",
            r.title,
            cov_cell,
        ]
        for c, val in enumerate(cells_row, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.alignment = WRAP_LEFT
        if r.req_id in orphans:
            for c in range(1, 5):
                ws.cell(row=row, column=c).fill = ORPHAN_FILL
        ws.row_dimensions[row].height = _row_height_for(r.title)
        row += 1

    # Summary block
    row += 1
    n_total = len(reqs_in_order)
    n_orphans = len(orphans)
    n_covered = n_total - n_orphans
    pct = (100.0 * n_covered / n_total) if n_total else 0.0
    summary = (
        f"Coverage summary: {n_covered}/{n_total} "
        f"requirements claimed by ≥ 1 flow ({pct:.1f}%); "
        f"{n_orphans} orphan requirement(s)."
    )
    cell = ws.cell(row=row, column=1, value=summary)
    cell.font = Font(bold=True, color="1F3A5F")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)


def _write_flows_sheet(wb, plan: Plan) -> None:
    """A second auxiliary sheet: one row per flow with its scope, equipment,
    categories, and covered req-IDs. Useful for QA leads selecting which
    flows to schedule for a given test cycle.
    """
    flows_with_reqs = plan.__dict__.get("_flows_with_reqs", [])
    ws = wb.create_sheet("Flows")
    widths = [14, 40, 60, 30, 30, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    headers = ["Flow ID", "Flow Name", "Summary", "Equipment",
               "Categories Tested", "Covered Req IDs"]
    for c, label in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 28

    row = 2
    # Always render the full catalog so reviewers can also see flows
    # that produced no rows (e.g. because no req matched the selector).
    # Coverage-driven flows (scale, upgrade, NETCONF, OTF, soak, access-
    # interface variants) intentionally have no requirement anchor —
    # they are test techniques applied broadly. Render them with a
    # distinct marker so reviewers don't read the empty cell as a gap.
    coverage_fill = PatternFill("solid", fgColor="DCE9F1")  # neutral blue
    active_ids = {f.id for f, _ in flows_with_reqs}
    for flow in EVPN_FLOWS:
        covered = next((c for f, c in flows_with_reqs if f.id == flow.id), [])
        if covered:
            ids = ", ".join(r.req_id for r in covered)
        elif flow.coverage_driven:
            ids = (
                "(coverage-driven flow — test technique applied broadly; "
                "no single requirement anchor)"
            )
        else:
            ids = "(no req matched — flow catalog gap)"
        cells_row = [
            flow.id, flow.name, flow.summary, flow.equipment,
            ", ".join(flow.categories), ids,
        ]
        for c, val in enumerate(cells_row, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.alignment = WRAP_LEFT
        if flow.id not in active_ids:
            # Genuinely inactive — flag orange.
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).fill = ORPHAN_FILL
        elif not covered and flow.coverage_driven:
            # Active but coverage-driven — neutral blue, not alarming.
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).fill = coverage_fill
        ws.row_dimensions[row].height = _row_height_for(flow.summary)
        row += 1


def _write_rfc_crosscheck_sheet(wb, plan: Plan) -> None:
    """Render the RFC Cross-Check sheet: which RFCs the SFS cites, and which
    of them were actually ingested into the engine.

    Aleksey Burger (2026-06-04) asked the engine to "cross check and alert if
    there are reference RFCs specified in the SFS which weren't ingested".
    Missing RFCs are highlighted in orange with the SFS context snippet so a
    reviewer can judge which carry feature-relevant mandates worth ingesting.
    """
    catalog = plan.__dict__.get("_catalog")
    cc = getattr(catalog, "rfc_crosscheck", None) if catalog else None
    if cc is None:
        return
    ws = wb.create_sheet("RFC Cross-Check")
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 90

    row = 1
    ws.cell(row=row, column=1,
            value="SFS-cited RFCs vs. ingested RFCs").font = (
        Font(bold=True, size=12, color="1F3A5F")
    )
    row += 1
    intro = ws.cell(
        row=row, column=1,
        value=("Every RFC the SFS text references, reconciled against the "
               "RFCs actually provided as engine inputs. Rows highlighted in "
               "orange are cited by the SFS but were NOT ingested — their "
               "mandates are invisible to the generated plan. Add them under "
               "references/<FEATURE>/, or confirm they carry no feature-"
               "relevant requirements. (RFC 7432 is matched by the ingested "
               "rfc7432bis revision; BCP-14 keyword RFCs 2119/8174 are "
               "ignored.)"),
    )
    intro.alignment = WRAP_LEFT
    intro.font = Font(italic=True, color="555555")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.row_dimensions[row].height = _row_height_for(intro.value, width=110)
    row += 2

    for c, label in enumerate(("RFC", "Status", "SFS context (first mention)"), 1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[row].height = 24
    row += 1

    missing = set(cc.missing)
    ingested = cc.ingested
    for num in sorted(cc.cited, key=int):
        if num in missing:
            status = "MISSING — not ingested"
        elif num in ingested:
            status = "ingested ✓"
        else:
            status = "ignored (BCP-14 keyword RFC)"
        cells_row = [f"RFC{num}", status, cc.cited[num]]
        for c, val in enumerate(cells_row, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.alignment = WRAP_LEFT
        if num in missing:
            for c in range(1, 4):
                ws.cell(row=row, column=c).fill = ORPHAN_FILL
        ws.row_dimensions[row].height = _row_height_for(cc.cited[num], width=90)
        row += 1

    row += 1
    summary = (
        f"Cross-check summary: SFS cites {len(cc.cited)} RFC(s); "
        f"{len(cc.covered)} ingested, {len(cc.missing)} missing."
    )
    cell = ws.cell(row=row, column=1, value=summary)
    cell.font = Font(bold=True, color="1F3A5F")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)


def write_xlsx(plan: Plan, output_path: str | Path,
               cli_doc_path: str | Path | None = None) -> Path:
    output_path = Path(output_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Plan Topics"

    # 9-column widths matched to DHCP-snoopy reference TP.
    # Topic | Action | Req ID | Expectation | Monitor | Equipment | Build | Results | Comment
    widths = [32, 60, 22, 60, 36, 28, 12, 14, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # The 9-column header opens the sheet at row 1 and is the only frozen
    # row. The Feature meta + concept catalog used to sit above it, so
    # freeze_panes locked ~55 tall rows at the top — taller than a screen,
    # which made the body impossible to scroll into view (client review
    # 2026-06-14). Both now live on a separate "Overview" sheet.
    row = 1
    for c, label in enumerate(HEADER_ROW, 1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[row].height = 36
    row += 1
    ws.freeze_panes = ws.cell(row=row, column=1).coordinate

    # ── Body: PlanRow blobs → AtomicRow stream (DHCP-snoopy shape) ────
    # The generator still emits multi-line Setup/Action/Verify PlanRow
    # blobs (so the AI-enrichment cache survives the shape change);
    # atomic_rows decomposes each blob into a banner + atomic action
    # rows that match the reference TP's per-action layout.
    flow_lookup = {f.id: f for f in EVPN_FLOWS}
    inherited_names = set()
    catalog = plan.__dict__.get("_catalog")
    if catalog is not None:
        inherited_names = set(catalog.inherited_cmd_names)

    # Build a req_by_id lookup so we can read each PlanRow's anchor
    # req kind (delta / overlay / pointer) — Yossi 2026-05-21 follow-up.
    req_by_id = {r.req_id: r for r in plan.requirements}

    # Group rows into the DHCP-snoopy section bands. We stable-sort a
    # local index copy (plan.rows is left untouched, so the Plan model and
    # determinism tests are unaffected): within a band the original order
    # is preserved — CLI commands keep catalog order, flows keep FLOW-NNN
    # order, RFC mandates keep section order. A dark section header is
    # written at each band transition; the existing light topic banners
    # render underneath, matching references/DHCP-snoopy_TP_with_PW.xlsx.
    indexed = sorted(
        enumerate(plan.rows),
        key=lambda t: (_section_rank(_section_for_row(t[1], flow_lookup)),
                       t[0]),
    )

    current_section: str | None = None
    last_topic: str | None = None
    for _orig_idx, r in indexed:
        section = _section_for_row(r, flow_lookup)
        if section != current_section:
            row = _write_section_header(ws, section, row)
            current_section = section
            last_topic = None  # force the first topic banner in each band
        provenance = _provenance_for_row(r, req_by_id, inherited_names)
        # Suppress banner if it would repeat the previous banner (e.g.
        # multiple PlanRows for the same CLI command).
        topic_now = (f"{r.flow_id} — {r.flow_name}" if r.flow_id
                      else (r.sub_category or r.category or ""))
        for ar in rows_for_plan_row(
            r, flow_lookup=flow_lookup,
            emit_banner=(topic_now != last_topic),
            provenance=provenance,
        ):
            row = _write_atomic_row(ws, ar, row)
        last_topic = topic_now

    # ── Overview sheet: feature meta + concept catalog ─────────────────
    # Relocated off the main "Test Plan Topics" sheet so the test-plan
    # body scrolls freely under a single frozen header row (client review
    # 2026-06-14).
    ov = wb.create_sheet("Overview")
    for col, w in (("A", 32), ("B", 28), ("C", 90)):
        ov.column_dimensions[col].width = w
    orow = 1
    ov.cell(row=orow, column=1, value=f"Feature: {plan.feature_name}").font = META_FONT
    orow += 1
    ov.cell(row=orow, column=1, value=f"Source: {plan.source_path}")
    orow += 1
    ov.cell(row=orow, column=1, value="Machine vendors").font = META_FONT
    ov.cell(row=orow, column=2, value=plan.machine_vendor)
    orow += 1
    ov.cell(row=orow, column=1, value="Machine types").font = META_FONT
    ov.cell(row=orow, column=2, value=plan.machine_types)
    orow += 1
    ov.cell(row=orow, column=1, value="IP versions").font = META_FONT
    ov.cell(row=orow, column=2, value=plan.ip_versions)
    orow += 1
    ov.cell(row=orow, column=1, value="Interfaces").font = META_FONT
    ov.cell(row=orow, column=2, value=plan.interfaces)
    orow += 1
    ov.cell(row=orow, column=1, value=f"Requirements found: {plan.n_requirements}").font = META_FONT
    orow += 1
    flows_with_reqs = plan.__dict__.get("_flows_with_reqs", [])
    n_anchored = sum(1 for f, c in flows_with_reqs if c)
    n_coverage = sum(1 for f, c in flows_with_reqs
                     if not c and f.coverage_driven)
    n_dormant = len(EVPN_FLOWS) - len(flows_with_reqs)
    flow_summary = (
        f"Flows: {n_anchored} requirement-anchored + "
        f"{n_coverage} coverage-driven (test technique, applies broadly) "
        f"= {n_anchored + n_coverage} active of {len(EVPN_FLOWS)} "
        f"catalogued"
    )
    if n_dormant:
        flow_summary += f"; {n_dormant} catalogued but inactive in this run"
    ov.cell(row=orow, column=1, value=flow_summary).font = META_FONT
    orow += 2

    overview_catalog = build_catalog(cli_doc_path)
    if overview_catalog:
        _write_catalog(ov, overview_catalog, orow)

    # ── Requirements traceability sheet ────────────────────────────────
    ws2 = wb.create_sheet("Requirements")
    ws2.append(["Req ID", "Section", "Title", "Description (truncated)"])
    for c in range(1, 5):
        cell = ws2.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 50
    ws2.column_dimensions["D"].width = 80
    for r in plan.requirements:
        ws2.append([r.req_id, r.section_number or "", r.title,
                    (r.description[:300] + "…")
                    if len(r.description) > 300 else r.description])

    _write_flows_sheet(wb, plan)
    _write_coverage_sheet(wb, plan)
    _write_rfc_crosscheck_sheet(wb, plan)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
