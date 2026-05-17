"""Planner tests — IR → Plan → xlsx (M1 deliverable verification)."""
from __future__ import annotations

from pathlib import Path

import openpyxl

from ate.parsers import parse
from ate.planner import generate_plan, generate_plan_to_xlsx
from ate.planner.categories import ALL_CATEGORIES
from ate.planner.extractor import extract_requirements

ROOT = Path(__file__).resolve().parents[1]
EVPN_SPEC = ROOT / "tests/corpus/tier_a/EVPN System Specification 1.00.docx"
RFC7432BIS = ROOT / "references" / "draft-ietf-bess-rfc7432bis-13.txt"
RFC9785 = ROOT / "references" / "rfc9785.txt"


# ─── Requirement extraction ─────────────────────────────────────────────────

def test_extracts_evpn_anchors_from_evpn_spec() -> None:
    d = parse(EVPN_SPEC)
    reqs = extract_requirements(d)
    assert len(reqs) >= 30, f"expected ≥30 requirements, got {len(reqs)}"
    # First anchor should be EVPNS-REQ#10
    assert reqs[0].req_id == "EVPNS-REQ#10"
    # Each requirement has a non-empty title
    assert all(r.title for r in reqs), "some requirements lack titles"


def test_extractor_dedupes_repeated_anchors() -> None:
    d = parse(EVPN_SPEC)
    reqs = extract_requirements(d)
    ids = [r.req_id for r in reqs]
    assert len(ids) == len(set(ids)), "extractor returned duplicate req_ids"


# ─── Plan model ─────────────────────────────────────────────────────────────

def test_plan_from_evpn_spec_uses_template_categories() -> None:
    plan = generate_plan(EVPN_SPEC)
    assert plan.feature_name, "feature_name must be set"
    assert plan.n_requirements >= 30
    # Every category referenced in plan rows must be one of the template's
    rows_categories = {r.category for r in plan.rows}
    for cat in rows_categories:
        assert cat in ALL_CATEGORIES, f"unexpected category {cat!r} in plan"
    # Tech-support always applies (in ALWAYS_CATEGORIES)
    assert "Tech-support" in rows_categories


def test_plan_rows_are_traced_to_a_requirement() -> None:
    plan = generate_plan(EVPN_SPEC)
    for r in plan.rows:
        assert r.sfs_requirement_id, f"row missing requirement_id: {r}"
        assert r.action_steps, f"row missing action_steps: {r}"
        assert r.expectation, f"row missing expectation: {r}"


def test_plan_applies_categories_per_flow_not_uniformly() -> None:
    """Each flow gets only the categories meaningful to its use case —
    not all 17. v1 produced 24 rows per req; the QA-respin pivot picks
    a per-flow subset (Basic / Packet / On-the-fly / …) per flow."""
    plan = generate_plan(EVPN_SPEC)
    rows_per_flow: dict[str, list[str]] = {}
    for r in plan.rows:
        if not r.flow_id:
            continue
        rows_per_flow.setdefault(r.flow_id, []).append(r.category)
    counts = sorted(len(v) for v in rows_per_flow.values())
    assert counts, "no flow rows emitted — flow catalog is not matching"
    assert counts[0] < counts[-1], (
        "all flows got the same row count — categories are not being "
        "filtered per flow"
    )


# ─── xlsx writer ────────────────────────────────────────────────────────────

def test_xlsx_is_written_and_readable(tmp_path: Path) -> None:
    out = tmp_path / "plan.xlsx"
    plan = generate_plan_to_xlsx(EVPN_SPEC, out)
    assert out.exists() and out.stat().st_size > 0
    wb = openpyxl.load_workbook(out)
    assert "Test Plan Topics" in wb.sheetnames
    assert "Requirements" in wb.sheetnames
    # Requirements sheet header + at least one data row per requirement
    ws = wb["Requirements"]
    assert ws.max_row >= plan.n_requirements + 1


def test_xlsx_columns_match_template_schema(tmp_path: Path) -> None:
    """Column header for the DHCP-snoopy 9-column shape (M1 client respin
    2026-05-17). `references/DHCP-snoopy_TP_with_PW.xlsx` is the visual
    target; columns mirror its layout so QA reads atomic actions per row
    instead of multi-line Setup/Action/Verify blobs."""
    out = tmp_path / "plan.xlsx"
    generate_plan_to_xlsx(EVPN_SPEC, out)
    wb = openpyxl.load_workbook(out)
    ws = wb["Test Plan Topics"]
    header_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Topic":
            header_row = r
            break
    assert header_row is not None, "Could not find 'Topic' header row"
    expected = [
        "Topic",
        "Action",
        "SFS / RFC Req ID",
        "Expectation",
        "Monitor (show / verify command)",
        "Test Equipment",
        "Build number",
        "Results (Pass\\Fail)",
        "Comment \\ Bug number",
    ]
    for c, exp in enumerate(expected, 1):
        assert ws.cell(row=header_row, column=c).value == exp, (
            f"column {c}: expected {exp!r}, got {ws.cell(row=header_row, column=c).value!r}"
        )


def test_xlsx_contains_evpn_anchors(tmp_path: Path) -> None:
    """Req ID column (col 3 in the DHCP-snoopy 9-col schema) lists every
    requirement that the atomic-action row claims. EVPNS-REQ#NN anchors
    must appear across the column's comma-joined values."""
    out = tmp_path / "plan.xlsx"
    generate_plan_to_xlsx(EVPN_SPEC, out)
    wb = openpyxl.load_workbook(out)
    ws = wb["Test Plan Topics"]
    seen_anchors: set[str] = set()
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=3).value
        if not isinstance(v, str):
            continue
        for tok in (s.strip() for s in v.split(",")):
            if tok.startswith("EVPNS-REQ#"):
                seen_anchors.add(tok)
    assert len(seen_anchors) >= 20, f"only {len(seen_anchors)} anchors in xlsx"


def test_plan_handles_doc_without_anchors(tmp_path: Path) -> None:
    """A doc with no EVPNS-REQ# anchors should still produce a non-empty plan."""
    txt = tmp_path / "no_anchors.txt"
    txt.write_text(
        "1.  Test Feature\n\n    This is a feature with no requirement IDs.\n"
    )
    plan = generate_plan(txt)
    assert plan.n_rows > 0
    # Synthetic placeholder requirement
    assert any(r.req_id == "(no-anchor)" for r in plan.requirements)


# ─── Determinism ────────────────────────────────────────────────────────────

def test_planner_is_deterministic() -> None:
    p1 = generate_plan(EVPN_SPEC)
    p2 = generate_plan(EVPN_SPEC)
    assert p1.model_dump_json() == p2.model_dump_json()


# ─── CLI integration ────────────────────────────────────────────────────────

def test_cli_plan_command_works(tmp_path: Path) -> None:
    from ate.cli import main as cli_main
    out = tmp_path / "plan.xlsx"
    rc = cli_main(["plan", str(EVPN_SPEC), "-o", str(out), "--summary"])
    assert rc == 0
    rc = cli_main(["plan", str(EVPN_SPEC), "-o", str(out)])
    assert rc == 0
    assert out.exists() and out.stat().st_size > 0


def test_cli_plan_handles_bad_input(tmp_path: Path, capsys) -> None:
    from ate.cli import main as cli_main
    bad = tmp_path / "nope.html"
    bad.write_bytes(b"<html>not supported</html>")
    rc = cli_main(["plan", str(bad), "-o", str(tmp_path / "x.xlsx")])
    assert rc == 1
    err = capsys.readouterr().err
    assert "UnsupportedFormatError" in err


# ─── RFC integration (M1 respin) ────────────────────────────────────────────

def test_plan_with_rfcs_adds_rfc_anchored_rows() -> None:
    """Generating with --rfc paths must surface RFC*-§N entries in the
    flow-row coverage lists alongside EVPNS-REQ# entries."""
    base = generate_plan(EVPN_SPEC, use_ai=False)
    enriched = generate_plan(EVPN_SPEC, use_ai=False,
                             rfc_paths=[RFC7432BIS, RFC9785])
    assert enriched.n_requirements > base.n_requirements
    assert enriched.n_rows >= base.n_rows
    coverage_tokens: set[str] = set()
    for r in enriched.rows:
        for cid in r.covered_req_ids:
            coverage_tokens.add(cid)
    assert any(a.startswith("EVPNS-REQ#") for a in coverage_tokens)
    assert any(a.startswith("RFC7432bis-§") for a in coverage_tokens)
    assert any(a.startswith("RFC9785-§") for a in coverage_tokens)


def test_plan_dedupes_overlapping_rfc_paths() -> None:
    """Passing the same RFC twice must not double-count requirements."""
    once = generate_plan(EVPN_SPEC, use_ai=False, rfc_paths=[RFC9785])
    twice = generate_plan(EVPN_SPEC, use_ai=False, rfc_paths=[RFC9785, RFC9785])
    assert once.n_requirements == twice.n_requirements
    assert once.n_rows == twice.n_rows


def test_xlsx_with_rfcs_contains_both_anchor_styles(tmp_path: Path) -> None:
    """Req ID column (col 3) carries comma-joined req-IDs the atomic row
    claims; both EVPNS-REQ# and RFC*-§ tokens must appear."""
    out = tmp_path / "plan_rfc.xlsx"
    generate_plan_to_xlsx(EVPN_SPEC, out, use_ai=False,
                          rfc_paths=[RFC7432BIS, RFC9785])
    wb = openpyxl.load_workbook(out)
    ws = wb["Test Plan Topics"]
    spec_anchors: set[str] = set()
    rfc_anchors: set[str] = set()
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=3).value
        if not isinstance(v, str):
            continue
        for tok in (s.strip() for s in v.split(",")):
            if tok.startswith("EVPNS-REQ#"):
                spec_anchors.add(tok)
            elif tok.startswith("RFC"):
                rfc_anchors.add(tok)
    assert len(spec_anchors) >= 20
    assert len(rfc_anchors) >= 10, f"only {len(rfc_anchors)} RFC anchors in xlsx"


def test_cli_plan_accepts_repeated_rfc_flag(tmp_path: Path) -> None:
    from ate.cli import main as cli_main
    out = tmp_path / "plan.xlsx"
    rc = cli_main([
        "plan", str(EVPN_SPEC), "-o", str(out),
        "--rfc", str(RFC7432BIS),
        "--rfc", str(RFC9785),
        "--no-ai",
    ])
    assert rc == 0
    wb = openpyxl.load_workbook(out)
    ws = wb["Requirements"]
    ids = {ws.cell(row=r, column=1).value
           for r in range(2, ws.max_row + 1)}
    assert any(isinstance(t, str) and t.startswith("RFC7432bis-§") for t in ids)
    assert any(isinstance(t, str) and t.startswith("RFC9785-§") for t in ids)


def test_rfc_only_flow_rows_skip_platform_categories() -> None:
    """A flow row whose covered_req_ids are entirely RFC-sourced cannot
    appear under platform-only categories (CLI, On-the-fly, Upgrade,
    Management). RFCs describe protocol behaviour, not vendor CLI."""
    plan = generate_plan(EVPN_SPEC, use_ai=False,
                         rfc_paths=[RFC7432BIS, RFC9785])
    forbidden = {"CLI configuration", "On The Fly changes",
                 "Upgrade", "Management"}
    offenders = []
    for r in plan.rows:
        if not r.covered_req_ids:
            continue
        if all(cid.startswith("RFC") for cid in r.covered_req_ids):
            if r.category in forbidden:
                offenders.append(r)
    assert not offenders, (
        f"{len(offenders)} RFC-only flow rows landed in platform-only "
        f"categories: "
        f"{[(o.flow_id, o.category) for o in offenders[:5]]}"
    )


def test_flow_row_count_is_bounded() -> None:
    """No single flow should produce more rows than its declared
    categories list — guards against accidental category duplication
    in the overlay path."""
    from collections import Counter
    plan = generate_plan(EVPN_SPEC, use_ai=False,
                         rfc_paths=[RFC7432BIS, RFC9785])
    counts = Counter(r.flow_id for r in plan.rows if r.flow_id)
    assert counts, "no flow rows were emitted"
    # A flow declares ~3-7 categories; allow some headroom but keep a
    # ceiling so a regression that fans rows out per req is caught.
    worst_id, worst_n = counts.most_common(1)[0]
    assert worst_n <= 10, (
        f"{worst_id} produced {worst_n} rows — flow-row fan-out regressed"
    )


def test_coverage_sheet_links_reqs_to_flows(tmp_path: Path) -> None:
    """Coverage sheet must list every spec / RFC requirement and the
    flows that exercise it (or '(orphan)' if no flow claims it)."""
    out = tmp_path / "plan_cov.xlsx"
    generate_plan_to_xlsx(EVPN_SPEC, out, use_ai=False,
                          rfc_paths=[RFC7432BIS, RFC9785])
    wb = openpyxl.load_workbook(out)
    assert "Coverage" in wb.sheetnames
    ws = wb["Coverage"]
    seen_reqs: set[str] = set()
    flow_links = 0
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        d = ws.cell(row=r, column=4).value
        if isinstance(a, str) and (a.startswith("EVPNS-REQ#") or a.startswith("RFC")):
            seen_reqs.add(a)
            if isinstance(d, str) and "FLOW-" in d:
                flow_links += 1
    assert len(seen_reqs) >= 30, f"coverage only lists {len(seen_reqs)} reqs"
    assert flow_links >= 10, (
        f"only {flow_links} reqs are linked to flows in Coverage sheet"
    )
